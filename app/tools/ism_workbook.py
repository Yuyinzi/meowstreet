from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook


@dataclass(frozen=True)
class RankingLayout:
    sheet: str
    header_row: int
    data_row: int
    industry_column: int
    first_status_column: int


def iso_date(value):
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return str(value)


def is_date(value):
    if isinstance(value, (date, datetime)):
        return True
    if not isinstance(value, str):
        return False
    try:
        datetime.fromisoformat(value)
    except ValueError:
        return False
    return True


def _parse_series_points(sheet, source):
    rows = list(sheet.iter_rows(min_row=2, min_col=1, max_col=2, values_only=True))
    if not rows:
        raise ValueError("sheet has no data rows")
    points = []
    seen_dates = set()
    for row_index, (date_value, level_value) in enumerate(rows, start=2):
        if not is_date(date_value):
            raise ValueError(f"row {row_index} has invalid date: {date_value!r}")
        if level_value in (None, ""):
            raise ValueError(f"row {row_index} has empty level value")
        date = iso_date(date_value)
        if date in seen_dates:
            raise ValueError(f"row {row_index} has duplicate date: {date!r}")
        seen_dates.add(date)
        points.append(
            {
                "date": date,
                "value": float(level_value),
                "source": source,
            }
        )
    return points


def parse_series_workbook(workbook_path, survey_label, series_config):
    path = Path(workbook_path)
    if not path.exists():
        raise ValueError(f"{survey_label} workbook is missing: {path}")
    workbook = load_workbook(path, read_only=True, data_only=True)
    results = []
    for series_id, config in series_config.items():
        sheet_name = config["sheet"]
        if sheet_name not in workbook.sheetnames:
            raise ValueError(f"{survey_label} sheet is missing: {sheet_name}")
        sheet = workbook[sheet_name]
        points = _parse_series_points(sheet, path.name)
        results.append(
            {
                "series": {
                    "series_id": series_id,
                    "title": config["title"],
                    "units": config["units"],
                    "source": path.name,
                },
                "points": points,
            }
        )
    return results


def _direction(value):
    values = {"Growth": "growth", "Contraction": "contraction", "Neutral": "neutral"}
    return values.get(value)


def parse_ranking_workbook(workbook_path, survey_type, survey_label, layout):
    path = Path(workbook_path)
    if not path.exists():
        raise ValueError(f"{survey_label} workbook is missing: {path}")
    workbook = load_workbook(path, read_only=True, data_only=True)
    if layout.sheet not in workbook.sheetnames:
        raise ValueError(f"{survey_label} sheet is missing: {layout.sheet}")
    sheet = workbook[layout.sheet]
    month_columns = []
    for column in range(layout.first_status_column, sheet.max_column + 1, 2):
        header = sheet.cell(layout.header_row, column).value
        if header is None:
            break
        if not is_date(header):
            raise ValueError(
                f"{survey_label} ranking column {column} header is not a date: {header!r}"
            )
        month_columns.append((column, iso_date(header)))
    if not month_columns:
        raise ValueError(
            f"{survey_label} ranking sheet has no date headers starting at column {layout.first_status_column}"
        )
    rows = []
    seen = set()
    known_industry_names = set()
    for row_index in range(layout.data_row, sheet.max_row + 1):
        industry = sheet.cell(row_index, layout.industry_column).value
        if not industry:
            continue
        name = str(industry).strip()
        for status_column, month in month_columns:
            raw_direction = sheet.cell(row_index, status_column).value
            rank = sheet.cell(row_index, status_column + 1).value
            direction = _direction(raw_direction)
            if direction is None:
                if isinstance(raw_direction, str):
                    raise ValueError(
                        f"{survey_label} ranking has unknown direction {raw_direction!r} for {name} in {month}"
                    )
                continue
            if rank in (None, ""):
                raise ValueError(
                    f"{survey_label} ranking has missing rank for {name} in {month}"
                )
            key = (survey_type, month, name)
            if key in seen:
                raise ValueError(
                    f"{survey_label} ranking has duplicate row for {name} in {month}"
                )
            seen.add(key)
            known_industry_names.add(name)
            rows.append(
                {
                    "survey_type": survey_type,
                    "date": month,
                    "industry": name,
                    "direction": direction,
                    "rank": int(rank),
                    "source": path.name,
                }
            )
    if not rows:
        raise ValueError(f"{survey_label} ranking sheet produced no valid rows")
    return rows
