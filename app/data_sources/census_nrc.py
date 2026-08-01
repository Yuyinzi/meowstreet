from datetime import datetime
from pathlib import Path

import openpyxl

from app.http_client import HttpClient


PERMIT_HISTORY_URL = "https://www.census.gov/construction/nrc/xls/permits_cust.xlsx"


def _find_seasonally_adjusted_sheet(workbook):
    for name in workbook.sheetnames:
        if "seasonally adjusted" in str(name).lower():
            return workbook[name]
    raise ValueError(
        "workbook has no seasonally adjusted sheet, "
        f"available sheets: {workbook.sheetnames}"
    )


def _header_row_index(ws):
    for index, row in enumerate(
        ws.iter_rows(min_row=1, max_row=min(5, ws.max_row), values_only=True)
    ):
        if row and row[0] and str(row[0]).strip().lower() == "month":
            return index
    raise ValueError("seasonally adjusted sheet has no Month header row")


def _is_numeric(value):
    if value is None:
        return False
    if isinstance(value, (int, float)):
        return True
    return False


def _is_data_row(row):
    if not row or not row[0]:
        return False
    first = row[0]
    if isinstance(first, datetime):
        return True
    if isinstance(first, str) and first.strip().lower().startswith(
        (
            "january",
            "february",
            "march",
            "april",
            "may",
            "june",
            "july",
            "august",
            "september",
            "october",
            "november",
            "december",
        )
    ):
        return False
    return False


def fetch_permits_workbook(destination, http_client=None):
    destination = Path(destination)
    destination.parent.mkdir(parents=True, exist_ok=True)
    client = http_client or HttpClient()
    response = client.request("GET", PERMIT_HISTORY_URL, timeout=60)
    destination.write_bytes(response.content)
    return destination


def parse_permits_workbook(workbook_path, release_date=None):
    workbook_path = Path(workbook_path)
    workbook = openpyxl.load_workbook(workbook_path, data_only=True)
    ws = _find_seasonally_adjusted_sheet(workbook)
    header_idx = _header_row_index(ws)
    data_start = header_idx + 1
    observations = []
    seen_dates = set()
    release_date_value = release_date or datetime.today().strftime("%Y-%m-%d")
    for row_idx, row in enumerate(
        ws.iter_rows(min_row=data_start + 1, values_only=True), start=data_start + 1
    ):
        if row is None:
            continue
        date_val = row[0] if len(row) > 0 else None
        value_val = row[1] if len(row) > 1 else None
        if date_val is None:
            continue
        if not isinstance(date_val, datetime):
            continue
        if not _is_numeric(value_val):
            continue
        if value_val < 0:
            raise ValueError(f"negative permits value at row {row_idx}: {value_val}")
        obs_date = date_val.strftime("%Y-%m-%d")
        if obs_date in seen_dates:
            raise ValueError(
                f"duplicate observation month at row {row_idx}: {obs_date}"
            )
        seen_dates.add(obs_date)
        observations.append(
            {
                "date": obs_date,
                "value": float(value_val),
                "source": "census.xlsx",
                "release_date": release_date_value,
                "revision_status": "official_current_history",
                "source_url": PERMIT_HISTORY_URL,
                "source_identifier": workbook_path.name,
            }
        )
    observations.sort(key=lambda o: o["date"])
    return {
        "series": {
            "series_id": "building_permits_saar",
            "title": "Building Permits SAAR",
            "units": "thousands_saar",
            "source": "Census New Residential Construction",
        },
        "observations": observations,
    }
