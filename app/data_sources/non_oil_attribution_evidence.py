import io
import json
import math
from html.parser import HTMLParser
from urllib.parse import urljoin
from urllib.parse import urlparse

from openpyxl import load_workbook

from app.http_client import HttpClient

VERSION = "non_oil_attribution_evidence_v1"

IWCC_SOURCE_NAME = "International Wrought Copper Council"
IWCC_SOURCE_URL = "http://www.coppercouncil.org/iwcc-statistics-and-data"
IWCC_WORKBOOK_ANCHOR = "Semis production and demand.xlsx"
IWCC_WORKSHEET_LABEL = "Prod-Dem Summary External"
IWCC_TOTAL_BLOCK_LABEL = "Total Copper"
IWCC_GLOBAL_ROW_LABEL = "World Total"
IWCC_FACTOR_LABELS = {"Production": "supply", "Demand": "demand"}
IWCC_PRODUCTION_COLUMN_RANGE = (2, 15)
IWCC_DEMAND_COLUMN_RANGE = (15, 28)
IWCC_GEOGRAPHY = "Global"
IWCC_UNITS = "t"

FAOSTAT_SOURCE_NAME = "Food and Agriculture Organization of the United Nations"
FAOSTAT_SOURCE_URL = "https://www.fao.org/faostat/en/#data/FO"
FAOSTAT_API_URL = "https://fenixservices.fao.org/faostat/api/v1/en/data/FO"
FAOSTAT_ITEM = "Sawnwood"
FAOSTAT_AREA = "World"
FAOSTAT_ELEMENT_FACTORS = {
    "Production": "supply",
    "Import Quantity": "trade",
    "Export Quantity": "trade",
}
_FAOSTAT_REQUIRED_ROW_FIELDS = ("item", "area", "year", "element", "value", "unit")

_DEFAULT_CLIENT = HttpClient()


class _AnchorCollector(HTMLParser):
    def __init__(self):
        super().__init__()
        self._anchors = []
        self._current_href = None
        self._text_parts = []

    def handle_starttag(self, tag, attrs):
        if tag == "a":
            self._current_href = dict(attrs).get("href")
            self._text_parts = []

    def handle_data(self, data):
        if self._current_href is not None:
            self._text_parts.append(data)

    def handle_endtag(self, tag):
        if tag == "a" and self._current_href is not None:
            self._anchors.append((self._current_href, "".join(self._text_parts)))
            self._current_href = None

    def anchors(self):
        return self._anchors


def _normalize_anchor(text):
    return " ".join(str(text).strip().lower().split())


def fetch_iwcc_copper_facts(http_client=None):
    client = http_client or _DEFAULT_CLIENT
    page_response = client.request("GET", IWCC_SOURCE_URL)
    workbook_url = _iwcc_workbook_url(page_response.content)
    workbook_response = client.request("GET", workbook_url)
    raw_facts = _parse_iwcc_workbook(workbook_response.content)
    return [normalize_non_oil_attribution_fact(raw_fact) for raw_fact in raw_facts]


def _iwcc_workbook_url(page_content):
    collector = _AnchorCollector()
    collector.feed(page_content.decode("utf-8", errors="replace"))
    workbook_anchor = next(
        (
            href
            for href, text in collector.anchors()
            if _normalize_anchor(text) == _normalize_anchor(IWCC_WORKBOOK_ANCHOR)
        ),
        None,
    )
    if workbook_anchor is None:
        raise ValueError("iwcc semis production and demand workbook anchor is missing")
    resolved = urljoin(IWCC_SOURCE_URL, workbook_anchor)
    if urlparse(resolved).netloc != urlparse(IWCC_SOURCE_URL).netloc:
        raise ValueError(
            "iwcc semis production and demand workbook is not on the iwcc origin"
        )
    return resolved


def _parse_iwcc_workbook(content):
    workbook = _load_iwcc_workbook(content)
    sheet = _iwcc_summary_sheet(workbook)
    block = _iwcc_total_copper_block(sheet)
    production_columns = _iwcc_year_columns(
        sheet, block["year_row"], IWCC_PRODUCTION_COLUMN_RANGE
    )
    demand_columns = _iwcc_year_columns(
        sheet, block["year_row"], IWCC_DEMAND_COLUMN_RANGE
    )
    if not production_columns or not demand_columns:
        raise ValueError(
            "iwcc semis production and demand workbook is missing year headers"
        )
    columns_by_factor = {
        "Production": production_columns,
        "Demand": demand_columns,
    }
    global_rows = _iwcc_global_rows(sheet, block)
    if not global_rows:
        raise ValueError(
            "iwcc semis production and demand workbook is missing global cells"
        )
    if len(global_rows) > 1:
        raise ValueError(
            "iwcc semis production and demand workbook has duplicate global rows"
        )
    newest_year = max(set(production_columns) & set(demand_columns))
    global_row = global_rows[0]
    return [
        _iwcc_raw_fact(
            factor_category,
            factor_label,
            newest_year,
            _iwcc_global_value(
                global_row[columns_by_factor[factor_label][newest_year] - 1],
                factor_category,
            ),
        )
        for factor_label, factor_category in IWCC_FACTOR_LABELS.items()
    ]


def _load_iwcc_workbook(content):
    try:
        return load_workbook(io.BytesIO(content), data_only=True)
    except Exception as exc:
        raise ValueError(
            "iwcc semis production and demand workbook is invalid"
        ) from exc


def _iwcc_summary_sheet(workbook):
    for sheet in workbook.worksheets:
        if _normalize_anchor(sheet.title) == _normalize_anchor(IWCC_WORKSHEET_LABEL):
            return sheet
    raise ValueError(
        "iwcc semis production and demand workbook is missing the summary sheet"
    )


def _iwcc_total_copper_block(sheet):
    factor_rows = [
        row_index
        for row_index in range(1, sheet.max_row + 1)
        if _normalize_anchor(sheet.cell(row_index, 2).value) == "production"
        and _normalize_anchor(sheet.cell(row_index, 15).value) == "demand"
    ]
    for row_index in factor_rows:
        label = sheet.cell(row_index - 1, 2).value
        if _normalize_anchor(label) == _normalize_anchor(IWCC_TOTAL_BLOCK_LABEL):
            return {
                "label_row": row_index - 1,
                "factor_row": row_index,
                "year_row": row_index + 1,
            }
    raise ValueError("iwcc semis production and demand workbook is missing a factor")


def _iwcc_year_columns(sheet, year_row, column_range):
    start, end = column_range
    result = {}
    for column in range(start, end):
        value = sheet.cell(year_row, column).value
        year = _as_year(value)
        if year is not None and 1000 <= year <= 9999:
            result[year] = column
    return result


def _as_year(value):
    try:
        return int(value)
    except (TypeError, ValueError):
        return None


def _iwcc_global_rows(sheet, block):
    rows = []
    for row_index in range(block["factor_row"] + 2, sheet.max_row + 1):
        if (
            _normalize_anchor(sheet.cell(row_index, 2).value) == "production"
            and _normalize_anchor(sheet.cell(row_index, 15).value) == "demand"
        ):
            break
        label = sheet.cell(row_index, 1).value
        if _normalize_anchor(label) == _normalize_anchor(IWCC_GLOBAL_ROW_LABEL):
            rows.append(
                [
                    sheet.cell(row_index, column).value
                    for column in range(1, sheet.max_column + 1)
                ]
            )
    return rows


def _iwcc_raw_fact(factor_category, factor_label, year, value):
    return {
        "commodity_id": "copper",
        "source_name": IWCC_SOURCE_NAME,
        "source_url": IWCC_SOURCE_URL,
        "factor_category": factor_category,
        "metric_name": factor_label,
        "geography": IWCC_GEOGRAPHY,
        "observation_date": f"{year}-12-31",
        "publication_date": None,
        "value": value,
        "units": IWCC_UNITS,
    }


def _iwcc_global_value(raw_value, factor_category):
    try:
        value = float(raw_value)
    except (TypeError, ValueError) as exc:
        raise ValueError(
            f"iwcc semis production and demand global {factor_category} value is not numeric"
        ) from exc
    if not math.isfinite(value):
        raise ValueError(
            f"iwcc semis production and demand global {factor_category} value is not numeric"
        )
    return value


def fetch_faostat_lumber_facts(http_client=None):
    client = http_client or _DEFAULT_CLIENT
    response = client.request("GET", FAOSTAT_API_URL)
    raw_facts = _faostat_raw_facts(_faostat_payload(response.content))
    return [normalize_non_oil_attribution_fact(raw_fact) for raw_fact in raw_facts]


def _faostat_payload(content):
    try:
        payload = json.loads(content)
    except ValueError as exc:
        raise ValueError(
            "faostat forestry production and trade payload is invalid"
        ) from exc
    if not isinstance(payload, dict) or not isinstance(payload.get("data"), list):
        raise ValueError("faostat forestry production and trade payload is invalid")
    return payload


def _faostat_raw_facts(payload):
    rows = payload["data"]
    candidates = [
        row
        for row in rows
        if isinstance(row, dict)
        and row.get("item") == FAOSTAT_ITEM
        and row.get("area") == FAOSTAT_AREA
    ]
    _require_faostat_source_fields(candidates)
    return [
        _faostat_raw_fact(element_label, candidates)
        for element_label in FAOSTAT_ELEMENT_FACTORS
    ]


def _require_faostat_source_fields(rows):
    for row in rows:
        if not isinstance(row, dict):
            raise ValueError("faostat forestry production and trade row is invalid")
        missing = [
            field
            for field in _FAOSTAT_REQUIRED_ROW_FIELDS
            if row.get(field) in (None, "")
        ]
        if missing:
            raise ValueError(
                "faostat forestry production and trade row is missing a required source field"
            )


def _faostat_raw_fact(element_label, candidates):
    element_rows = [row for row in candidates if row.get("element") == element_label]
    if not element_rows:
        raise ValueError(
            f"faostat forestry production and trade is missing {element_label} facts for {FAOSTAT_ITEM} in {FAOSTAT_AREA}"
        )
    newest_row = max(element_rows, key=lambda row: _faostat_year(row))
    year = _faostat_year(newest_row)
    value = _faostat_value(newest_row, element_label)
    return {
        "commodity_id": "lumber",
        "source_name": FAOSTAT_SOURCE_NAME,
        "source_url": FAOSTAT_SOURCE_URL,
        "factor_category": FAOSTAT_ELEMENT_FACTORS[element_label],
        "metric_name": element_label,
        "geography": newest_row["area"],
        "observation_date": f"{year}-12-31",
        "publication_date": None,
        "value": value,
        "units": newest_row["unit"],
    }


def _faostat_year(row):
    try:
        return int(row["year"])
    except (TypeError, ValueError) as exc:
        raise ValueError(
            "faostat forestry production and trade year is invalid"
        ) from exc


def _faostat_value(row, element_label):
    try:
        value = float(row["value"])
    except (TypeError, ValueError) as exc:
        raise ValueError(
            f"faostat forestry production and trade {element_label} value is not numeric"
        ) from exc
    if not math.isfinite(value):
        raise ValueError(
            f"faostat forestry production and trade {element_label} value is not numeric"
        )
    return value


def normalize_non_oil_attribution_fact(raw_fact):
    required = (
        "commodity_id",
        "source_name",
        "source_url",
        "factor_category",
        "metric_name",
        "geography",
        "observation_date",
        "value",
        "units",
    )
    if any(raw_fact.get(key) in (None, "") for key in required):
        raise ValueError(" non-oil attribution fact has a required field missing")
    try:
        value = float(raw_fact["value"])
    except (TypeError, ValueError) as exc:
        raise ValueError(" non-oil attribution fact value is invalid") from exc
    if not math.isfinite(value):
        raise ValueError(" non-oil attribution fact value is invalid")
    return {
        **raw_fact,
        "method_version": VERSION,
        "value": value,
        "status": "available",
    }
