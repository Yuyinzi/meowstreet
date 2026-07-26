from datetime import datetime
from pathlib import Path

import openpyxl
import pytest

from app.data_sources import census_nrc
from app.db import macro_indicators
from app.services import housing_permits_import


def _census_xlsx_bytes():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Seasonally Adjusted"
    ws.append(["Month", "United States"])
    ws.append([datetime(2026, 4, 1), 1423])
    ws.append([datetime(2026, 5, 1), 1413])
    from io import BytesIO

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def write_census_permits_workbook(xlsx_path, observations):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Seasonally Adjusted"
    ws.append(["Month", "United States"])
    for obs_date, value in observations:
        ws.append([obs_date, value])
    wb.save(xlsx_path)


class FakeResponse:
    def __init__(self, content):
        self._content = content

    def read(self):
        return self._content

    def __enter__(self):
        return self

    def __exit__(self, *args):
        pass


class TestCensusSource:
    def test_parse_permits_workbook_reads_seasonally_adjusted_us_total(self, tmp_path):
        workbook_path = tmp_path / "permits_cust.xlsx"
        write_census_permits_workbook(
            workbook_path,
            [(datetime(2026, 4, 1), 1423), (datetime(2026, 5, 1), 1413)],
        )
        payload = census_nrc.parse_permits_workbook(
            workbook_path, release_date="2026-06-16"
        )
        assert payload["series"]["series_id"] == "building_permits_saar"
        assert payload["observations"][-1]["date"] == "2026-05-01"
        assert payload["observations"][-1]["value"] == 1413.0
        assert (
            payload["observations"][-1]["source_url"] == census_nrc.PERMIT_HISTORY_URL
        )

    def test_fetch_permits_workbook_writes_official_response(
        self, monkeypatch, tmp_path
    ):
        monkeypatch.setattr(
            census_nrc,
            "urlopen",
            lambda request, timeout: FakeResponse(_census_xlsx_bytes()),
        )
        destination = census_nrc.fetch_permits_workbook(tmp_path / "permits_cust.xlsx")
        assert destination.read_bytes() == _census_xlsx_bytes()

    def test_parse_permits_workbook_rejects_missing_seasonally_adjusted_sheet(
        self, tmp_path
    ):
        workbook_path = tmp_path / "permits_cust.xlsx"
        wb = openpyxl.Workbook()
        wb.active.title = "Not Adjusted"
        wb.save(workbook_path)
        with pytest.raises(ValueError, match="seasonally adjusted"):
            census_nrc.parse_permits_workbook(workbook_path)

    def test_parse_permits_workbook_rejects_duplicate_months(self, tmp_path):
        workbook_path = tmp_path / "permits_cust.xlsx"
        write_census_permits_workbook(
            workbook_path,
            [(datetime(2026, 4, 1), 1423), (datetime(2026, 4, 1), 1425)],
        )
        with pytest.raises(ValueError, match="duplicate"):
            census_nrc.parse_permits_workbook(workbook_path)

    def test_parse_permits_workbook_rejects_negative_value(self, tmp_path):
        workbook_path = tmp_path / "permits_cust.xlsx"
        write_census_permits_workbook(workbook_path, [(datetime(2026, 4, 1), -100)])
        with pytest.raises(ValueError, match="negative"):
            census_nrc.parse_permits_workbook(workbook_path)


class TestHousingPermitsImport:
    def test_import_cached_workbook_writes_observations_with_metadata(self, tmp_path):
        con = macro_indicators.connect(tmp_path / "market.sqlite")
        workbook_path = tmp_path / "permits_cust.xlsx"
        write_census_permits_workbook(
            workbook_path,
            [(datetime(2026, 4, 1), 1423), (datetime(2026, 5, 1), 1413)],
        )
        count = housing_permits_import.import_cached_official_workbook(
            con,
            workbook_path,
            release_date="2026-06-16",
        )
        assert count == 2
        observations = macro_indicators.load_macro_indicator_observations(
            con, "building_permits_saar"
        )
        assert len(observations) == 2
        assert observations[-1]["value"] == 1413.0
        assert observations[-1]["release_date"] == "2026-06-16"

    def test_import_cached_workbook_overwrites_revised_month(self, tmp_path):
        con = macro_indicators.connect(tmp_path / "market.sqlite")
        path_v1 = tmp_path / "v1.xlsx"
        write_census_permits_workbook(path_v1, [(datetime(2026, 5, 1), 1413)])
        housing_permits_import.import_cached_official_workbook(
            con, path_v1, release_date="2026-06-16"
        )
        path_v2 = tmp_path / "v2.xlsx"
        write_census_permits_workbook(path_v2, [(datetime(2026, 5, 1), 1418)])
        count = housing_permits_import.import_cached_official_workbook(
            con, path_v2, release_date="2026-07-17"
        )
        assert count == 1
        observations = macro_indicators.load_macro_indicator_observations(
            con, "building_permits_saar"
        )
        assert observations[0]["value"] == 1418.0

    def test_fetch_official_workbook_calls_census_fetch(self, monkeypatch, tmp_path):
        cache_path = tmp_path / "cached.xlsx"
        fetch_called = []

        def fake_fetch(destination):
            fetch_called.append(destination)
            Path(destination).write_bytes(_census_xlsx_bytes())
            return Path(destination)

        monkeypatch.setattr(census_nrc, "fetch_permits_workbook", fake_fetch)
        result = housing_permits_import.fetch_official_workbook(cache_path)
        assert str(result).endswith(".xlsx")
        assert len(fetch_called) == 1

    def test_refresh_official_history_fetches_then_imports(self, monkeypatch, tmp_path):
        con = macro_indicators.connect(tmp_path / "market.sqlite")
        cache_path = tmp_path / "cached.xlsx"
        Path(cache_path).write_bytes(_census_xlsx_bytes())
        fetch_called = []

        def fake_fetch(destination):
            fetch_called.append(str(destination))
            return Path(destination)

        monkeypatch.setattr(
            housing_permits_import, "fetch_official_workbook", fake_fetch
        )
        count = housing_permits_import.refresh_official_history(
            con, cache_path, release_date="2026-06-16"
        )
        assert count == 2
        assert len(fetch_called) == 1
