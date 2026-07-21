from datetime import datetime

import pytest
from openpyxl import Workbook

from app.tools import ism_workbook


def test_parse_series_workbook_reads_configured_sheet(tmp_path):
    path = tmp_path / "survey.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Activity"
    sheet.append(["Date", "Activity"])
    sheet.append([datetime(2026, 5, 1), 54.2])
    workbook.save(path)

    result = ism_workbook.parse_series_workbook(
        path,
        "ism services",
        {
            "ism_services_business_activity": {
                "sheet": "Activity",
                "title": "ISM Services Business Activity",
                "units": "index",
            }
        },
    )

    assert result[0]["points"] == [
        {"date": "2026-05-01", "value": 54.2, "source": "survey.xlsx"}
    ]


def test_parse_series_workbook_skips_blank_rows(tmp_path):
    path = tmp_path / "survey.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "NMI"
    sheet.append(["Date", "NMI"])
    sheet.append([datetime(2026, 5, 1), 54.2])
    sheet.append([datetime(2026, 6, 1), 55.1])
    sheet.append([None, None])
    sheet.append([None, None])
    workbook.save(path)

    result = ism_workbook.parse_series_workbook(
        path,
        "ism services",
        {"ism_services_pmi": {"sheet": "NMI", "title": "PMI", "units": "index"}},
    )

    assert len(result[0]["points"]) == 2
    assert result[0]["points"][1] == {
        "date": "2026-06-01",
        "value": 55.1,
        "source": "survey.xlsx",
    }


def test_parse_ranking_workbook_respects_end_row(tmp_path):
    path = tmp_path / "rankings.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sectors"
    sheet.append([None])
    sheet.append([None])
    sheet.append(["Sectors", datetime(2026, 6, 1)])
    sheet.append([None])
    sheet.append([None])
    sheet.append(["Construction", "Growth", 1])
    sheet.append(["Retail Trade", "Contraction", -1])
    sheet.append([None, None, None])
    sheet.append(["Business Activity", datetime(2026, 6, 1)])
    workbook.save(path)

    layout = ism_workbook.RankingLayout(
        sheet="Sectors",
        header_row=3,
        data_row=6,
        industry_column=1,
        first_status_column=2,
        end_row=7,
    )

    rows = ism_workbook.parse_ranking_workbook(path, "services", "test", layout)

    assert len(rows) == 2
    assert rows[0]["industry"] == "Construction"
    assert rows[1]["industry"] == "Retail Trade"


def test_parse_series_workbook_rejects_entirely_blank_sheet(tmp_path):
    path = tmp_path / "survey.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "NMI"
    sheet.append(["Date", "NMI"])
    sheet.append([None, None])
    sheet.append([None, None])
    workbook.save(path)

    with pytest.raises(ValueError, match="no valid data points"):
        ism_workbook.parse_series_workbook(
            path,
            "ism services",
            {"ism_services_pmi": {"sheet": "NMI", "title": "PMI", "units": "index"}},
        )


def test_parse_series_workbook_rejects_missing_sheet(tmp_path):
    path = tmp_path / "survey.xlsx"
    Workbook().save(path)

    with pytest.raises(ValueError, match="ism services sheet is missing: Activity"):
        ism_workbook.parse_series_workbook(
            path,
            "ism services",
            {"activity": {"sheet": "Activity", "title": "Activity", "units": "index"}},
        )
