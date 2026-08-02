import io
from urllib.parse import unquote

import httpx
import pytest
from openpyxl import Workbook

from app.data_sources import non_oil_attribution_evidence as evidence
from app.http_client import HttpClient

_IWCC_PAGE_URL = "http://www.coppercouncil.org/iwcc-statistics-and-data"
_IWCC_WORKBOOK_URL = (
    "http://www.coppercouncil.org/files/Semis production and demand.xlsx"
)
_IWCC_PAGE_HTML = (
    "<html><body><ul>"
    '<li><a href="/files/Semis production and demand.xlsx">Semis production and demand.xlsx</a></li>'
    "</ul></body></html>"
)


def valid_fact():
    return {
        "commodity_id": "copper",
        "source_name": "International Wrought Copper Council",
        "source_url": "http://www.coppercouncil.org/iwcc-statistics-and-data",
        "factor_category": "supply",
        "metric_name": "Production",
        "geography": "Global",
        "observation_date": "2024-12-31",
        "publication_date": None,
        "value": 24100.0,
        "units": "t",
    }


_IWCC_YEARS = list(range(2012, 2025))
_IWCC_PRODUCTION_START_COL = 2
_IWCC_DEMAND_START_COL = 15


def _iwcc_real_layout_workbook(
    production_2024=24100.0,
    demand_2024=23900.0,
    factor_label="Total Copper",
    include_demand=True,
    duplicate_global_rows=False,
):
    wb = Workbook()
    ws = wb.active
    ws.title = "Prod-Dem Summary External"
    ws.cell(1, _IWCC_PRODUCTION_START_COL, factor_label)
    ws.cell(2, _IWCC_PRODUCTION_START_COL, "Production")
    if include_demand:
        ws.cell(2, _IWCC_DEMAND_START_COL, "Demand")
    for idx, year in enumerate(_IWCC_YEARS):
        ws.cell(3, _IWCC_PRODUCTION_START_COL + idx, year)
        if include_demand:
            ws.cell(3, _IWCC_DEMAND_START_COL + idx, year)
    ws.cell(4, 1, "EU27+UK")
    for idx, year in enumerate(_IWCC_YEARS):
        ws.cell(4, _IWCC_PRODUCTION_START_COL + idx, 1000 + year)
        if include_demand:
            ws.cell(4, _IWCC_DEMAND_START_COL + idx, 2000 + year)
    ws.cell(5, 1, "World Total")
    for idx, year in enumerate(_IWCC_YEARS):
        ws.cell(5, _IWCC_PRODUCTION_START_COL + idx, 21000 + year)
        if include_demand:
            ws.cell(5, _IWCC_DEMAND_START_COL + idx, 22000 + year)
    ws.cell(5, _IWCC_PRODUCTION_START_COL + len(_IWCC_YEARS) - 1, production_2024)
    if include_demand:
        ws.cell(5, _IWCC_DEMAND_START_COL + len(_IWCC_YEARS) - 1, demand_2024)
    if duplicate_global_rows:
        ws.cell(6, 1, "World Total")
        for idx, year in enumerate(_IWCC_YEARS):
            ws.cell(6, _IWCC_PRODUCTION_START_COL + idx, 21000 + year + 100)
            if include_demand:
                ws.cell(6, _IWCC_DEMAND_START_COL + idx, 22000 + year + 100)
    stream = io.BytesIO()
    wb.save(stream)
    return stream.getvalue()


def _iwcc_workbook_bytes():
    return _iwcc_real_layout_workbook()


def iwcc_handler(request):
    if unquote(str(request.url)) == _IWCC_PAGE_URL:
        return httpx.Response(200, text=_IWCC_PAGE_HTML)
    if unquote(str(request.url)) == _IWCC_WORKBOOK_URL:
        return httpx.Response(200, content=_iwcc_workbook_bytes())
    return httpx.Response(404)


def _faostat_payload():
    return {
        "data": [
            {
                "area": "World",
                "item": "Sawnwood",
                "year": 2022,
                "element": "Production",
                "unit": "m3",
                "value": "251000000",
            },
            {
                "area": "World",
                "item": "Sawnwood",
                "year": 2023,
                "element": "Production",
                "unit": "m3",
                "value": "252000000",
            },
            {
                "area": "World",
                "item": "Sawnwood",
                "year": 2024,
                "element": "Production",
                "unit": "m3",
                "value": "254000000",
            },
            {
                "area": "World",
                "item": "Sawnwood",
                "year": 2023,
                "element": "Import Quantity",
                "unit": "m3",
                "value": "112000000",
            },
            {
                "area": "World",
                "item": "Sawnwood",
                "year": 2024,
                "element": "Import Quantity",
                "unit": "m3",
                "value": "114000000",
            },
            {
                "area": "World",
                "item": "Sawnwood",
                "year": 2024,
                "element": "Export Quantity",
                "unit": "m3",
                "value": "116000000",
            },
            {
                "area": "Europe",
                "item": "Sawnwood",
                "year": 2024,
                "element": "Production",
                "unit": "m3",
                "value": "50000000",
            },
        ]
    }


def faostat_handler(request):
    if str(request.url) == evidence.FAOSTAT_API_URL:
        return httpx.Response(200, json=_faostat_payload())
    return httpx.Response(404)


def test_fetch_iwcc_returns_latest_global_production_and_demand():
    facts = evidence.fetch_iwcc_copper_facts(
        HttpClient(transport=httpx.MockTransport(iwcc_handler))
    )
    assert {(row["factor_category"], row["observation_date"]) for row in facts} == {
        ("supply", "2024-12-31"),
        ("demand", "2024-12-31"),
    }


def test_fetch_iwcc_parses_real_single_sheet_dual_block_layout():
    content = _iwcc_real_layout_workbook(production_2024=24100.0, demand_2024=23900.0)

    def handler(request):
        if unquote(str(request.url)) == _IWCC_PAGE_URL:
            return httpx.Response(200, text=_IWCC_PAGE_HTML)
        return httpx.Response(200, content=content)

    facts = evidence.fetch_iwcc_copper_facts(
        HttpClient(transport=httpx.MockTransport(handler))
    )
    assert {row["factor_category"] for row in facts} == {"supply", "demand"}
    assert all(row["observation_date"] == "2024-12-31" for row in facts)
    by_factor = {row["factor_category"]: row["value"] for row in facts}
    assert by_factor == {"supply": 24100.0, "demand": 23900.0}
    assert all(row["geography"] == "Global" for row in facts)


def test_fetch_iwcc_uses_total_copper_block_when_multiple_product_blocks_present():
    wb = Workbook()
    ws = wb.active
    ws.title = "Prod-Dem Summary External"
    for block_start, label, production_2024, demand_2024 in (
        (1, "Copper Wire Rod", 7000.0, 7100.0),
        (17, "Total Copper", 24100.0, 23900.0),
        (33, "Alloy Wire", 5000.0, 5100.0),
    ):
        ws.cell(block_start, 2, label)
        ws.cell(block_start + 1, 2, "Production")
        ws.cell(block_start + 1, 15, "Demand")
        for idx, year in enumerate(_IWCC_YEARS):
            ws.cell(block_start + 2, 2 + idx, year)
            ws.cell(block_start + 2, 15 + idx, year)
        ws.cell(block_start + 3, 1, "EU27+UK")
        for idx, year in enumerate(_IWCC_YEARS):
            ws.cell(block_start + 3, 2 + idx, 1000 + year)
            ws.cell(block_start + 3, 15 + idx, 2000 + year)
        ws.cell(block_start + 4, 1, "World Total")
        for idx, year in enumerate(_IWCC_YEARS):
            ws.cell(block_start + 4, 2 + idx, 10000 + year)
            ws.cell(block_start + 4, 15 + idx, 11000 + year)
        ws.cell(block_start + 4, 2 + len(_IWCC_YEARS) - 1, production_2024)
        ws.cell(block_start + 4, 15 + len(_IWCC_YEARS) - 1, demand_2024)
    stream = io.BytesIO()
    wb.save(stream)
    content = stream.getvalue()

    def handler(request):
        if unquote(str(request.url)) == _IWCC_PAGE_URL:
            return httpx.Response(200, text=_IWCC_PAGE_HTML)
        return httpx.Response(200, content=content)

    facts = evidence.fetch_iwcc_copper_facts(
        HttpClient(transport=httpx.MockTransport(handler))
    )
    by_factor = {row["factor_category"]: row["value"] for row in facts}
    assert by_factor == {"supply": 24100.0, "demand": 23900.0}
    assert {row["metric_name"] for row in facts} == {"Production", "Demand"}


def test_fetch_faostat_returns_production_import_and_export_with_source_fields():
    facts = evidence.fetch_faostat_lumber_facts(
        HttpClient(transport=httpx.MockTransport(faostat_handler))
    )
    assert {row["metric_name"] for row in facts} == {
        "Production",
        "Import Quantity",
        "Export Quantity",
    }
    assert all(
        row["units"] and row["geography"] and row["observation_date"] for row in facts
    )


def test_faostat_lumber_selection_is_fixed_sawnwood_world():
    assert evidence.FAOSTAT_ITEM == "Sawnwood"
    assert evidence.FAOSTAT_AREA == "World"
    facts = evidence.fetch_faostat_lumber_facts(
        HttpClient(transport=httpx.MockTransport(faostat_handler))
    )
    assert all(row["geography"] == "World" for row in facts)
    assert all(row["commodity_id"] == "lumber" for row in facts)


def test_normalize_rejects_missing_period_geography_unit_factor_or_numeric_value():
    for field, value in [
        ("observation_date", None),
        ("geography", ""),
        ("units", ""),
        ("factor_category", None),
        ("value", "x"),
    ]:
        raw = valid_fact()
        raw[field] = value
        with pytest.raises(ValueError, match=" non-oil attribution fact"):
            evidence.normalize_non_oil_attribution_fact(raw)


def test_fetch_iwcc_requests_source_page_then_same_origin_workbook():
    requested = []

    def handler(request):
        requested.append(unquote(str(request.url)))
        if requested[-1] == _IWCC_PAGE_URL:
            return httpx.Response(200, text=_IWCC_PAGE_HTML)
        return httpx.Response(200, content=_iwcc_workbook_bytes())

    facts = evidence.fetch_iwcc_copper_facts(
        HttpClient(transport=httpx.MockTransport(handler))
    )
    assert requested == [_IWCC_PAGE_URL, _IWCC_WORKBOOK_URL]
    assert len(facts) == 2


def test_fetch_faostat_requests_forestry_production_trade_surface():
    requested = []

    def handler(request):
        requested.append(str(request.url))
        return httpx.Response(200, json=_faostat_payload())

    evidence.fetch_faostat_lumber_facts(
        HttpClient(transport=httpx.MockTransport(handler))
    )
    assert requested == [evidence.FAOSTAT_API_URL]


def test_fetch_iwcc_rejects_missing_semis_workbook_anchor():
    def handler(request):
        return httpx.Response(200, text="<html><body>no workbook links</body></html>")

    with pytest.raises(
        ValueError, match="iwcc semis production and demand workbook anchor is missing"
    ):
        evidence.fetch_iwcc_copper_facts(
            HttpClient(transport=httpx.MockTransport(handler))
        )


def test_fetch_iwcc_rejects_workbook_missing_a_factor():
    content = _iwcc_real_layout_workbook(include_demand=False)

    def handler(request):
        if unquote(str(request.url)) == _IWCC_PAGE_URL:
            return httpx.Response(200, text=_IWCC_PAGE_HTML)
        return httpx.Response(200, content=content)

    with pytest.raises(
        ValueError,
        match="iwcc semis production and demand workbook is missing a factor",
    ):
        evidence.fetch_iwcc_copper_facts(
            HttpClient(transport=httpx.MockTransport(handler))
        )


def test_fetch_iwcc_rejects_non_numeric_global_cell():
    content = _iwcc_real_layout_workbook(production_2024="x")

    def handler(request):
        if unquote(str(request.url)) == _IWCC_PAGE_URL:
            return httpx.Response(200, text=_IWCC_PAGE_HTML)
        return httpx.Response(200, content=content)

    with pytest.raises(
        ValueError,
        match="iwcc semis production and demand global supply value is not numeric",
    ):
        evidence.fetch_iwcc_copper_facts(
            HttpClient(transport=httpx.MockTransport(handler))
        )


def test_fetch_iwcc_rejects_duplicate_global_rows():
    content = _iwcc_real_layout_workbook(duplicate_global_rows=True)

    def handler(request):
        if unquote(str(request.url)) == _IWCC_PAGE_URL:
            return httpx.Response(200, text=_IWCC_PAGE_HTML)
        return httpx.Response(200, content=content)

    with pytest.raises(
        ValueError,
        match="iwcc semis production and demand workbook has duplicate global rows",
    ):
        evidence.fetch_iwcc_copper_facts(
            HttpClient(transport=httpx.MockTransport(handler))
        )


def test_fetch_iwcc_raises_on_non_200_source_page():
    with pytest.raises(httpx.HTTPStatusError):
        evidence.fetch_iwcc_copper_facts(
            HttpClient(
                transport=httpx.MockTransport(lambda request: httpx.Response(404))
            )
        )


def test_fetch_faostat_rejects_row_missing_required_source_field():
    payload = _faostat_payload()
    del payload["data"][0]["unit"]

    def handler(request):
        return httpx.Response(200, json=payload)

    with pytest.raises(
        ValueError,
        match="faostat forestry production and trade row is missing a required source field",
    ):
        evidence.fetch_faostat_lumber_facts(
            HttpClient(transport=httpx.MockTransport(handler))
        )


def test_fetch_faostat_ignores_malformed_non_candidate_row():
    payload = _faostat_payload()
    payload["data"].append(
        {"area": "World", "item": "Wood Fuel", "year": 2024, "element": "Production"}
    )

    def handler(request):
        return httpx.Response(200, json=payload)

    facts = evidence.fetch_faostat_lumber_facts(
        HttpClient(transport=httpx.MockTransport(handler))
    )
    assert {row["metric_name"] for row in facts} == {
        "Production",
        "Import Quantity",
        "Export Quantity",
    }


def test_fetch_faostat_rejects_non_numeric_selected_value():
    payload = _faostat_payload()
    for row in payload["data"]:
        if (
            row["item"] == "Sawnwood"
            and row["area"] == "World"
            and row["element"] == "Production"
            and row["year"] == 2024
        ):
            row["value"] = "x"

    def handler(request):
        return httpx.Response(200, json=payload)

    with pytest.raises(
        ValueError,
        match="faostat forestry production and trade Production value is not numeric",
    ):
        evidence.fetch_faostat_lumber_facts(
            HttpClient(transport=httpx.MockTransport(handler))
        )


def test_fetch_faostat_rejects_missing_export_quantity():
    payload = _faostat_payload()
    payload["data"] = [
        row for row in payload["data"] if row["element"] != "Export Quantity"
    ]

    def handler(request):
        return httpx.Response(200, json=payload)

    with pytest.raises(
        ValueError,
        match="faostat forestry production and trade is missing Export Quantity facts",
    ):
        evidence.fetch_faostat_lumber_facts(
            HttpClient(transport=httpx.MockTransport(handler))
        )


def test_fetch_faostat_raises_on_non_200_surface():
    with pytest.raises(httpx.HTTPStatusError):
        evidence.fetch_faostat_lumber_facts(
            HttpClient(
                transport=httpx.MockTransport(lambda request: httpx.Response(404))
            )
        )


def test_fetch_facts_include_full_source_contract_fields():
    fetchers = [
        (
            evidence.fetch_iwcc_copper_facts,
            HttpClient(transport=httpx.MockTransport(iwcc_handler)),
            "http://www.coppercouncil.org/iwcc-statistics-and-data",
        ),
        (
            evidence.fetch_faostat_lumber_facts,
            HttpClient(transport=httpx.MockTransport(faostat_handler)),
            "https://www.fao.org/faostat/en/#data/FO",
        ),
    ]
    for fetch, client, source_url in fetchers:
        for fact in fetch(client):
            assert fact["method_version"] == "non_oil_attribution_evidence_v1"
            assert fact["source_url"] == source_url
            assert fact["factor_category"] in {"supply", "demand", "trade"}
            assert fact["value"] >= 0
            assert fact["status"] == "available"
            for field in (
                "commodity_id",
                "source_name",
                "metric_name",
                "geography",
                "observation_date",
                "units",
            ):
                assert fact.get(field) not in (None, "")


def test_fetch_facts_are_deterministic_for_same_source_fixture():
    fetchers = [
        (evidence.fetch_iwcc_copper_facts, iwcc_handler),
        (evidence.fetch_faostat_lumber_facts, faostat_handler),
    ]
    for fetch, handler in fetchers:
        first = fetch(HttpClient(transport=httpx.MockTransport(handler)))
        second = fetch(HttpClient(transport=httpx.MockTransport(handler)))
        assert first == second
