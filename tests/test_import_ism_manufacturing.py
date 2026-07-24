from datetime import datetime

import pytest
from openpyxl import Workbook, load_workbook

from app.db import growth_cycle
from app.db import macro_indicators
from app.db import us_rates_liquidity
from scripts import import_ism_manufacturing


ISM_TEST_SHEETS = [
    "PMI",
    "New Orders",
    "Production",
    "Employment",
    "Deliveries",
    "Inventories",
    "Customer Inventories",
    "Prices",
    "Order Backlog",
    "Exports",
    "Imports",
]


def write_ism_workbook(path):
    workbook = Workbook()
    for index, name in enumerate(ISM_TEST_SHEETS):
        if index == 0:
            sheet = workbook.active
            sheet.title = name
        else:
            sheet = workbook.create_sheet(title=name)
        sheet.append(["Date", name])
        sheet.append([datetime(2026, 1, 1), 50.0 + index])
        sheet.append([datetime(2026, 2, 1), 51.0 + index])
        sheet.append([datetime(2026, 3, 1), 52.0 + index])
        sheet.append([datetime(2026, 4, 1), 53.0 + index])
    sectors = workbook.create_sheet(title="Sectors")
    sectors.append([None] * 18)
    sectors.append([None] * 18)
    sectors.append(
        [
            None,
            "ISM Manufacturing",
            datetime(2026, 5, 1),
            None,
            datetime(2026, 6, 1),
            None,
        ]
    )
    sectors.append([None] * 18)
    sectors.append([None] * 18)
    sectors.append(
        [
            None,
            "Computer & Electronic Products",
            "Growth",
            8,
            "Growth",
            16,
        ]
    )
    sectors.append(
        [
            None,
            "Furniture & Related Products",
            "Growth",
            7,
            "Contraction",
            -1,
        ]
    )
    sectors.append(
        [
            None,
            "Machinery",
            "Contraction",
            -1,
            "Contraction",
            -2,
        ]
    )
    workbook.save(path)


def test_parse_workbook_reads_sector_rankings(tmp_path):
    workbook_path = tmp_path / "ISM_Manufacturing_Index.xlsx"
    write_ism_workbook(workbook_path)

    rankings = import_ism_manufacturing.parse_sector_rankings(workbook_path)

    assert rankings == [
        {
            "date": "2026-05-01",
            "industry": "Computer & Electronic Products",
            "direction": "growth",
            "rank": 8,
            "source": "ISM_Manufacturing_Index.xlsx",
        },
        {
            "date": "2026-06-01",
            "industry": "Computer & Electronic Products",
            "direction": "growth",
            "rank": 16,
            "source": "ISM_Manufacturing_Index.xlsx",
        },
        {
            "date": "2026-05-01",
            "industry": "Furniture & Related Products",
            "direction": "growth",
            "rank": 7,
            "source": "ISM_Manufacturing_Index.xlsx",
        },
        {
            "date": "2026-06-01",
            "industry": "Furniture & Related Products",
            "direction": "contraction",
            "rank": -1,
            "source": "ISM_Manufacturing_Index.xlsx",
        },
        {
            "date": "2026-05-01",
            "industry": "Machinery",
            "direction": "contraction",
            "rank": -1,
            "source": "ISM_Manufacturing_Index.xlsx",
        },
        {
            "date": "2026-06-01",
            "industry": "Machinery",
            "direction": "contraction",
            "rank": -2,
            "source": "ISM_Manufacturing_Index.xlsx",
        },
    ]


def test_parse_workbook_reads_all_ism_sheets(tmp_path):
    workbook_path = tmp_path / "ISM_Manufacturing_Index.xlsx"
    write_ism_workbook(workbook_path)

    results = import_ism_manufacturing.parse_workbook(workbook_path)

    assert len(results) == 11
    series_ids = [r["series"]["series_id"] for r in results]
    assert series_ids == list(import_ism_manufacturing.SERIES_CONFIG)
    pmi = next(
        r for r in results if r["series"]["series_id"] == "ism_manufacturing_pmi"
    )
    assert pmi["series"] == {
        "series_id": "ism_manufacturing_pmi",
        "title": "ISM Manufacturing PMI",
        "units": "index",
        "source": "ISM_Manufacturing_Index.xlsx",
    }
    assert pmi["points"] == [
        {"date": "2026-01-01", "value": 50.0, "source": "ISM_Manufacturing_Index.xlsx"},
        {"date": "2026-02-01", "value": 51.0, "source": "ISM_Manufacturing_Index.xlsx"},
        {"date": "2026-03-01", "value": 52.0, "source": "ISM_Manufacturing_Index.xlsx"},
        {"date": "2026-04-01", "value": 53.0, "source": "ISM_Manufacturing_Index.xlsx"},
    ]
    imports_series = next(
        r for r in results if r["series"]["series_id"] == "ism_manufacturing_imports"
    )
    assert imports_series["points"][0]["value"] == 50.0 + 10


def test_parse_workbook_rejects_missing_workbook(tmp_path):
    with pytest.raises(ValueError, match="ism manufacturing workbook is missing"):
        import_ism_manufacturing.parse_workbook(tmp_path / "missing.xlsx")


def test_parse_workbook_rejects_missing_sheet(tmp_path):
    workbook_path = tmp_path / "bad.xlsx"
    workbook = Workbook()
    workbook.active.title = "Wrong Sheet"
    workbook.save(workbook_path)

    with pytest.raises(ValueError, match="ism manufacturing sheet is missing"):
        import_ism_manufacturing.parse_workbook(workbook_path)


def test_import_workbook_saves_all_series_to_macro_indicator_tables(tmp_path):
    db_path = tmp_path / "market_data.sqlite"
    workbook_path = tmp_path / "ISM_Manufacturing_Index.xlsx"
    write_ism_workbook(workbook_path)
    con = us_rates_liquidity.connect(db_path)
    growth_cycle.init_db(con)

    inserted = import_ism_manufacturing.import_workbook(con, workbook_path)

    assert set(inserted) == set(import_ism_manufacturing.SERIES_CONFIG) | {
        "ism_industry_rankings"
    }
    for sid in import_ism_manufacturing.SERIES_CONFIG:
        assert inserted[sid] == 4
    assert inserted["ism_industry_rankings"] == 6
    pmi_points = macro_indicators.load_macro_indicator_points(
        con, "ism_manufacturing_pmi"
    )
    assert pmi_points == [
        {"date": "2026-01-01", "value": 50.0, "source": "ISM_Manufacturing_Index.xlsx"},
        {"date": "2026-02-01", "value": 51.0, "source": "ISM_Manufacturing_Index.xlsx"},
        {"date": "2026-03-01", "value": 52.0, "source": "ISM_Manufacturing_Index.xlsx"},
        {"date": "2026-04-01", "value": 53.0, "source": "ISM_Manufacturing_Index.xlsx"},
    ]
    rankings = growth_cycle.load_latest_ism_industry_rankings(con)
    assert rankings == [
        {
            "date": "2026-06-01",
            "industry": "Computer & Electronic Products",
            "direction": "growth",
            "rank": 16,
            "source": "ISM_Manufacturing_Index.xlsx",
        },
        {
            "date": "2026-06-01",
            "industry": "Furniture & Related Products",
            "direction": "contraction",
            "rank": -1,
            "source": "ISM_Manufacturing_Index.xlsx",
        },
        {
            "date": "2026-06-01",
            "industry": "Machinery",
            "direction": "contraction",
            "rank": -2,
            "source": "ISM_Manufacturing_Index.xlsx",
        },
    ]


def test_import_workbook_preserves_official_report_points(tmp_path):
    db_path = tmp_path / "market_data.sqlite"
    workbook_path = tmp_path / "ISM_Manufacturing_Index.xlsx"
    write_ism_workbook(workbook_path)
    con = us_rates_liquidity.connect(db_path)
    growth_cycle.init_db(con)
    macro_indicators.merge_macro_indicator_points(
        con,
        {
            "series_id": "ism_manufacturing_pmi",
            "title": "ISM Manufacturing PMI",
            "units": "index",
            "source": "ISM AI extraction",
        },
        [
            {
                "date": "2026-05-01",
                "value": 54.0,
                "source": "ISM AI extraction",
            }
        ],
    )

    import_ism_manufacturing.import_workbook(con, workbook_path)

    pmi_points = macro_indicators.load_macro_indicator_points(
        con, "ism_manufacturing_pmi"
    )
    assert pmi_points[-1] == {
        "date": "2026-05-01",
        "value": 54.0,
        "source": "ISM AI extraction",
    }


def test_import_workbook_validation_failure_does_not_change_series(tmp_path):
    db_path = tmp_path / "market_data.sqlite"
    workbook_path = tmp_path / "ISM_Manufacturing_Index.xlsx"
    write_ism_workbook(workbook_path)
    workbook = load_workbook(workbook_path)
    sectors = workbook["Sectors"]
    sectors.append(
        [
            None,
            "Machinery",
            "Contraction",
            -2,
            "Contraction",
            -3,
        ]
    )
    workbook.save(workbook_path)
    con = us_rates_liquidity.connect(db_path)
    growth_cycle.init_db(con)
    macro_indicators.merge_macro_indicator_points(
        con,
        {
            "series_id": "ism_manufacturing_pmi",
            "title": "ISM Manufacturing PMI",
            "units": "index",
            "source": "ISM AI extraction",
        },
        [
            {
                "date": "2026-01-01",
                "value": 99.0,
                "source": "ISM AI extraction",
            }
        ],
    )

    with pytest.raises(ValueError, match="duplicate row"):
        import_ism_manufacturing.import_workbook(con, workbook_path)

    pmi_points = macro_indicators.load_macro_indicator_points(
        con, "ism_manufacturing_pmi"
    )
    assert pmi_points == [
        {
            "date": "2026-01-01",
            "value": 99.0,
            "source": "ISM AI extraction",
        }
    ]


def test_parse_sector_rankings_raises_on_duplicate_row(tmp_path):
    workbook_path = tmp_path / "ISM_Manufacturing_Index.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sectors"
    sheet.append([None] * 18)
    sheet.append([None] * 18)
    sheet.append(
        [
            None,
            "ISM Manufacturing",
            datetime(2026, 6, 1),
            None,
            datetime(2026, 7, 1),
            None,
        ]
    )
    sheet.append([None] * 18)
    sheet.append([None] * 18)
    sheet.append([None, "Machinery", "Contraction", -2, "Contraction", -3])
    sheet.append([None, "Machinery", "Contraction", -2, "Contraction", -3])
    workbook.save(workbook_path)

    with pytest.raises(ValueError, match="duplicate row for Machinery in 2026-06-01"):
        import_ism_manufacturing.parse_sector_rankings(workbook_path)


def test_main_imports_ism_via_cli(tmp_path, capsys):
    db_path = tmp_path / "market_data.sqlite"
    workbook_path = tmp_path / "ISM_Manufacturing_Index.xlsx"
    write_ism_workbook(workbook_path)

    exit_code = import_ism_manufacturing.main(
        ["--db-path", str(db_path), "--workbook-path", str(workbook_path)]
    )

    assert exit_code == 0
    out = capsys.readouterr().out
    assert "ism_manufacturing_pmi: 4" in out
    assert "ism_manufacturing_imports: 4" in out
    for series_id in import_ism_manufacturing.SERIES_CONFIG:
        assert f"{series_id}: 4" in out
    assert "ism_industry_rankings: 6" in out
