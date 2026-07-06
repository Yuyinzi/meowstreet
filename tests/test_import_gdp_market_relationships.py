import datetime

import pytest
from openpyxl import Workbook

from app.db import gdp_market_relationships
from scripts import import_benchmark_market_data
from scripts import import_gdp_market_relationships


def make_workbook(path):
    wb = Workbook()
    wb.remove(wb.active)

    corr = wb.create_sheet("S&P500_USGDP Correlation")
    corr.append(
        [
            "Date",
            "S&P500",
            "US Real GDP",
            "S&P500 YoY",
            "GDP YoY",
            "Rolling 10yr-Correlation: NO LAG",
            "S&P500 YoY",
            "GDP YoY",
            "Rolling 10yr-Correlation: 3-MO LAG",
            "S&P500 YoY",
            "GDP YoY",
            "Rolling 10yr-Correlation: 6-MO LAG",
            "S&P500 YoY",
            "GDP YoY",
            "Rolling 10yr-Correlation: 9-MO LAG",
            "S&P500 YoY",
            "GDP YoY",
            "Rolling 10yr-Correlation: 12-MO LAG",
        ]
    )
    corr.append(
        [
            None,
            None,
            None,
            "No Lag",
            None,
            None,
            None,
            "S&P 500 3-Month Lag",
            None,
            None,
            "S&P 500 6-Month Lag",
            None,
            None,
            "S&P 500 9-Month Lag",
            None,
            None,
            "S&P 500 12-Month Lag",
            None,
        ]
    )
    corr.append(
        [
            datetime.datetime(2020, 6, 30),
            3100.29,
            17302.511,
            0.0538,
            -0.0903,
            0.1729,
            -0.0881,
            -0.0903,
            0.2602,
            0.2887,
            -0.0903,
            -0.0944,
            0.3188,
            -0.0903,
            -0.1211,
            0.3499,
            -0.0903,
            -0.1512,
        ]
    )

    quad = wb.create_sheet("S&P500_US_Quadnomial")
    quad.append(
        [
            "Date",
            "S&P500 6-MO LAG",
            None,
            "US Real GDP",
            "S&P",
            "GDP",
            "Total",
            "0,0",
            "1,1",
            "0,1",
            "1,0",
        ]
    )
    quad.append(
        [
            datetime.datetime(2020, 9, 30),
            2584.59,
            "2020 Q3",
            18596.521,
            0,
            1,
            1,
            None,
            None,
            1,
            None,
        ]
    )

    wb.save(path)


def test_parse_configured_relationship_extracts_lags_and_quad_rows(tmp_path):
    workbook_path = tmp_path / "GDP_Correlations.xlsx"
    make_workbook(workbook_path)

    relationship, lag_rows, quad_rows = (
        import_gdp_market_relationships.parse_relationship(
            workbook_path,
            import_gdp_market_relationships.GDP_RELATIONSHIP_SHEETS[0],
        )
    )

    assert relationship["relationship_id"] == "us_sp500_gdp"
    assert relationship["primary_lag_months"] == 6
    assert [row["lag_months"] for row in lag_rows] == [0, 3, 6, 9, 12]
    assert lag_rows[-1]["rolling_correlation"] == -0.1512
    assert quad_rows[0]["quad_case"] == "0,1"


def test_import_workbook_saves_relationship_data(tmp_path):
    db_path = tmp_path / "gdp.sqlite"
    workbook_path = tmp_path / "GDP_Correlations.xlsx"
    make_workbook(workbook_path)
    con = gdp_market_relationships.connect(db_path)

    inserted, errors = import_gdp_market_relationships.import_workbook(
        con,
        workbook_path,
        configs=[import_gdp_market_relationships.GDP_RELATIONSHIP_SHEETS[0]],
    )

    assert inserted["us_sp500_gdp"] == {"lag_rows": 5, "quad_rows": 1}
    assert errors == {}
    assert (
        gdp_market_relationships.load_relationships(con)[0]["relationship_id"]
        == "us_sp500_gdp"
    )


def test_video_03_china_quadnomial_sheet_uses_non_china_gdp_series():
    workbook_path = import_gdp_market_relationships.DEFAULT_WORKBOOK_PATH

    _, _, europe_quad_rows = import_gdp_market_relationships.parse_relationship(
        workbook_path,
        import_gdp_market_relationships.GDP_RELATIONSHIP_SHEETS[1],
    )
    _, _, china_quad_rows = import_gdp_market_relationships.parse_relationship(
        workbook_path,
        import_gdp_market_relationships.GDP_RELATIONSHIP_SHEETS[3],
    )
    china_correlation_sheet = import_benchmark_market_data.load_workbook_sheet(
        workbook_path,
        import_gdp_market_relationships.GDP_RELATIONSHIP_SHEETS[3]["correlation_sheet"],
        data_only=True,
    )
    china_first_gdp_level = china_correlation_sheet.cell(5, 3).value

    assert china_quad_rows[0]["gdp_level"] != china_first_gdp_level
    assert china_quad_rows[0]["gdp_level"] == europe_quad_rows[0]["gdp_level"]
    assert china_quad_rows[1]["gdp_level"] == europe_quad_rows[1]["gdp_level"]


def test_parse_fred_gdp_csv_converts_quarter_start_to_quarter_end(tmp_path):
    csv_path = tmp_path / "GDPC1.csv"
    csv_path.write_text(
        "observation_date,GDPC1\n2025-01-01,23548.210\n2025-04-01,23770.976\n",
        encoding="utf-8",
    )

    rows = import_gdp_market_relationships.parse_fred_gdp_csv(csv_path)

    assert rows == {
        "2025-03-31": 23548.210,
        "2025-06-30": 23770.976,
    }


def test_parse_fred_sp500_csv_uses_last_non_empty_close_per_quarter(tmp_path):
    csv_path = tmp_path / "SP500.csv"
    csv_path.write_text(
        "observation_date,SP500\n"
        "2025-03-28,5600.00\n"
        "2025-03-31,\n"
        "2025-04-01,5625.00\n"
        "2025-06-30,5900.00\n",
        encoding="utf-8",
    )

    rows = import_gdp_market_relationships.parse_fred_sp500_csv(csv_path)

    assert rows == {
        "2025-03-31": 5600.00,
        "2025-06-30": 5900.00,
    }


def test_import_us_csv_merge_uses_existing_db_history_for_rolling_correlation(tmp_path):
    db_path = tmp_path / "gdp.sqlite"
    gdp_csv = tmp_path / "GDPC1.csv"
    sp500_csv = tmp_path / "SP500.csv"
    con = gdp_market_relationships.connect(db_path)
    relationship = {
        "relationship_id": "us_sp500_gdp",
        "title": "S&P 500 vs US GDP",
        "region": "US",
        "economy": "US GDP",
        "index_name": "S&P 500",
        "primary_lag_months": 6,
        "correlation_window_years": 10,
        "source_workbook": "seed",
        "source_sheet": "seed",
    }
    seed_lag_rows = [
        {
            "date": f"{year}-{month_day}",
            "lag_months": 0,
            "index_yoy": 0.01 + index / 1000,
            "gdp_yoy": 0.02 + index / 1000,
            "rolling_correlation": None,
            "source_workbook": "seed",
            "source_sheet": "seed",
        }
        for index, (year, month_day) in enumerate(
            [
                (year, month_day)
                for year in range(2014, 2024)
                for month_day in ["03-31", "06-30", "09-30", "12-31"]
            ]
        )
    ]
    gdp_market_relationships.replace_relationship_data(
        con,
        relationship,
        seed_lag_rows,
        [],
    )
    gdp_csv.write_text(
        "observation_date,GDPC1\n2023-01-01,100\n2024-01-01,110\n",
        encoding="utf-8",
    )
    sp500_csv.write_text(
        "observation_date,SP500\n2023-03-31,200\n2024-03-31,230\n",
        encoding="utf-8",
    )

    saved = import_gdp_market_relationships.import_us_csv_merge(con, gdp_csv, sp500_csv)

    rows = gdp_market_relationships.load_lag_rows(con, "us_sp500_gdp")
    latest = [
        row for row in rows if row["date"] == "2024-03-31" and row["lag_months"] == 0
    ][0]

    assert saved["lag_rows"] > 0
    assert latest["index_yoy"] == pytest.approx(0.15)
    assert latest["gdp_yoy"] == pytest.approx(0.10)
    assert latest["rolling_correlation"] is not None
    assert latest["source_sheet"] == "computed"


def test_main_us_csv_merge_prints_before_after_summary(monkeypatch, capsys):
    class FakeConnection:
        def close(self):
            pass

    detail_by_call = [
        {
            "latest": {
                "primary_lag_date": "2020-09-30",
                "quadnomial_date": "2020-09-30",
                "rolling_index_gdp_correlation": 0.02,
                "average_10y_correlation": 0.56,
                "index_yoy": -0.08,
                "gdp_yoy": -0.02,
                "quadnomial_current_case": "0,1",
            },
            "same_direction_pct": 69.04,
            "method_explainable_pct": 94.66,
            "relationship_signal_usability": "GDP relationship usable",
            "macro_relationship_confidence": "high",
        },
        {
            "latest": {
                "primary_lag_date": "2026-03-31",
                "quadnomial_date": "2026-03-31",
                "rolling_index_gdp_correlation": 0.18,
                "average_10y_correlation": 0.53,
                "index_yoy": 0.16,
                "gdp_yoy": 0.03,
                "quadnomial_current_case": "1,1",
            },
            "same_direction_pct": 68.98,
            "method_explainable_pct": 94.39,
            "relationship_signal_usability": "GDP relationship usable",
            "macro_relationship_confidence": "high",
        },
    ]
    detail_calls = []

    def fake_connect():
        return FakeConnection()

    def fake_import_us_csv_merge(con):
        assert con is not None
        return {"lag_rows": 175, "quad_rows": 37}

    def fake_load_relationship(con, relationship_id):
        assert con is not None
        assert relationship_id == "us_sp500_gdp"
        return {
            "relationship_id": "us_sp500_gdp",
            "title": "S&P 500 vs US GDP",
            "region": "US",
            "economy": "US GDP",
            "index_name": "S&P 500",
            "primary_lag_months": 6,
            "correlation_window_years": 10,
        }

    def fake_build_summary(con, relationship_id):
        assert con is not None
        assert relationship_id == "us_sp500_gdp"
        detail = detail_by_call[len(detail_calls)]
        detail_calls.append(detail)
        return detail

    monkeypatch.setattr(
        import_gdp_market_relationships.gdp_market_relationships,
        "connect",
        fake_connect,
    )
    monkeypatch.setattr(
        import_gdp_market_relationships, "import_us_csv_merge", fake_import_us_csv_merge
    )
    monkeypatch.setattr(
        import_gdp_market_relationships, "_load_relationship", fake_load_relationship
    )
    monkeypatch.setattr(
        import_gdp_market_relationships,
        "_build_relationship_summary",
        fake_build_summary,
    )
    monkeypatch.setattr(
        import_gdp_market_relationships.sys,
        "argv",
        ["import_gdp_market_relationships.py", "--us-csv-merge"],
    )

    import_gdp_market_relationships.main()

    output = capsys.readouterr().out

    assert "us_sp500_gdp: 175 csv lag rows, 37 csv quad rows merged" in output
    assert "latest metric comparison for us_sp500_gdp" in output
    assert "primary_lag_date: 2020-09-30 -> 2026-03-31" in output
    assert "quadnomial_current_case: 0,1 -> 1,1" in output
    assert "rolling_index_gdp_correlation: 0.02 -> 0.18" in output
    assert "same_direction_pct: 69.04 -> 68.98" in output


def test_fetch_fred_csvs_uses_shared_fred_client(tmp_path, monkeypatch):
    fetched = []

    class FakeFredClient:
        def __init__(self, cache_dir):
            assert cache_dir == tmp_path

        def fetch_csv(self, series_id):
            fetched.append(series_id)
            return tmp_path / f"{series_id}.csv"

    monkeypatch.setattr(import_gdp_market_relationships, "FredClient", FakeFredClient)

    result = import_gdp_market_relationships.fetch_fred_csvs(
        fred_dir=tmp_path,
    )

    assert fetched == ["GDPC1", "SP500"]
    assert result == {
        "gdp_csv": str(tmp_path / "GDPC1.csv"),
        "sp500_csv": str(tmp_path / "SP500.csv"),
    }
