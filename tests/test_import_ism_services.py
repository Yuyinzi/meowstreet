from datetime import datetime

import pytest
from openpyxl import Workbook

from app.db import ism_surveys, macro_indicators, us_rates_liquidity
from scripts import import_ism_services


def _write_base_sheets(workbook):
    for index, name in enumerate(
        ["NMI", "Business Activity", "New Orders", "Order Backlog"]
    ):
        sheet = workbook.active if index == 0 else workbook.create_sheet()
        sheet.title = name
        sheet.append(["Date", name])
        sheet.append([datetime(2026, 5, 1), 51.0 + index])
    sectors = workbook.create_sheet("Sectors")
    sectors.append([None] * 6)
    sectors.append([None] * 6)
    sectors.append(["ISM Services", datetime(2026, 5, 1), None])
    sectors.append([None] * 6)
    sectors.append([None] * 6)
    sectors.append(["Construction", "Growth", 1])
    comments = workbook.create_sheet("Industry Comments")
    comments.append(["Sector", "Date", "Comment"])


def write_services_workbook(path):
    workbook = Workbook()
    _write_base_sheets(workbook)
    comments = workbook["Industry Comments"]
    comments.append(["Construction", datetime(2026, 5, 1), "Pipeline remains healthy."])
    workbook.save(path)


def test_import_workbook_saves_four_series_and_services_rankings(tmp_path):
    workbook_path = tmp_path / "services.xlsx"
    db_path = tmp_path / "market.sqlite"
    write_services_workbook(workbook_path)
    con = us_rates_liquidity.connect(db_path)

    inserted = import_ism_services.import_workbook(con, workbook_path)

    assert inserted["ism_services_pmi"] == 1
    assert inserted["ism_services_industry_rankings"] == 1
    assert inserted["ism_services_industry_comments"] == 1


def test_import_workbook_rejects_duplicate_after_normalization(tmp_path):
    workbook_path = tmp_path / "services.xlsx"
    db_path = tmp_path / "market.sqlite"
    workbook = Workbook()
    _write_base_sheets(workbook)
    workbook["Industry Comments"].append(
        ["Construction", datetime(2026, 5, 1), "Pipeline remains healthy."]
    )
    workbook["Sectors"].append(["Retail  Trade", "Growth", 1])
    workbook["Sectors"].append(["Retail Trade", "Growth", 2])
    workbook.save(workbook_path)

    con = import_ism_services.us_rates_liquidity.connect(db_path)
    try:
        import_ism_services.import_workbook(con, workbook_path)
    except ValueError as exc:
        assert "duplicate" in str(exc).lower()
    else:
        pytest.fail("expected ValueError for duplicate after normalization")
    finally:
        con.close()


def test_import_workbook_rolls_back_on_failure(tmp_path):
    class _FailOnComments:
        def __init__(self, con):
            self.con = con

        def __getattr__(self, name):
            return getattr(self.con, name)

        def execute(self, sql, *args):
            if "ism_industry_comments" in sql:
                raise RuntimeError("simulated write failure")
            return self.con.execute(sql, *args)

    workbook_path = tmp_path / "services.xlsx"
    db_path = tmp_path / "market.sqlite"
    workbook = Workbook()
    _write_base_sheets(workbook)
    workbook["Industry Comments"].append(
        ["Construction", datetime(2026, 5, 1), "Pipeline remains healthy."]
    )
    workbook.save(workbook_path)

    con = us_rates_liquidity.connect(db_path)
    fail_con = _FailOnComments(con)
    try:
        import_ism_services.import_workbook(fail_con, workbook_path)
    except RuntimeError:
        pass
    else:
        pytest.fail("expected RuntimeError")
    finally:
        con.close()

    series_ids = [
        "ism_services_pmi",
        "ism_services_business_activity",
        "ism_services_new_orders",
        "ism_services_order_backlog",
    ]
    check_con = us_rates_liquidity.connect(db_path)
    try:
        for sid in series_ids:
            points = macro_indicators.load_macro_indicator_points(check_con, sid)
            assert len(points) == 0, f"{sid} not empty after rollback"
        rankings = ism_surveys.load_industry_rankings(check_con, "services")
        assert len(rankings) == 0
        comments = ism_surveys.load_industry_comments(check_con, "services")
        assert len(comments) == 0
    finally:
        check_con.close()


def test_import_real_workbook(tmp_path):
    real_path = import_ism_services.DEFAULT_WORKBOOK_PATH
    if not real_path.exists():
        pytest.skip(f"real workbook not found: {real_path}")
    db_path = tmp_path / "market.sqlite"
    con = us_rates_liquidity.connect(db_path)
    try:
        inserted = import_ism_services.import_workbook(con, real_path)
    finally:
        con.close()

    assert inserted["ism_services_pmi"] > 0
    assert inserted["ism_services_business_activity"] > 0
    assert inserted["ism_services_new_orders"] > 0
    assert inserted["ism_services_order_backlog"] > 0
    assert inserted["ism_services_industry_rankings"] > 0
    assert inserted["ism_services_industry_comments"] > 0


def test_parse_industry_comments_rejects_orphan_continuation_row(tmp_path):
    path = tmp_path / "services.xlsx"
    workbook = Workbook()
    _write_base_sheets(workbook)
    workbook["Industry Comments"].append(
        [None, datetime(2026, 5, 1), "Orphan comment without industry"]
    )
    workbook.save(path)

    with pytest.raises(ValueError, match="no preceding industry"):
        import_ism_services.parse_industry_comments(path)


def test_parse_industry_comments_allows_trailing_empty_rows(tmp_path):
    path = tmp_path / "services.xlsx"
    workbook = Workbook()
    _write_base_sheets(workbook)
    workbook["Industry Comments"].append(
        ["Construction", datetime(2026, 5, 1), "Pipeline remains healthy."]
    )
    workbook["Industry Comments"].append([None, None, None])
    workbook.save(path)

    rows = import_ism_services.parse_industry_comments(path)
    assert len(rows) == 1


def test_parse_industry_comments_skips_empty_comments(tmp_path):
    path = tmp_path / "services.xlsx"
    workbook = Workbook()
    _write_base_sheets(workbook)
    comments = workbook["Industry Comments"]
    comments.append(["Construction", datetime(2026, 5, 1), "Real comment"])
    comments.append([None, datetime(2026, 4, 1), ""])
    comments.append([None, datetime(2026, 4, 1), None])
    workbook.save(path)

    rows = import_ism_services.parse_industry_comments(path)

    assert len(rows) == 1
    assert rows[0]["comment_text"] == "Real comment"


def test_parse_industry_comments_rejects_industry_without_date_or_comment(tmp_path):
    path = tmp_path / "services.xlsx"
    workbook = Workbook()
    _write_base_sheets(workbook)
    workbook["Industry Comments"].append(["Construction", None, None])
    workbook.save(path)

    with pytest.raises(ValueError, match="no date or comment"):
        import_ism_services.parse_industry_comments(path)


def test_parse_industry_comments_rejects_comment_without_date(tmp_path):
    path = tmp_path / "services.xlsx"
    workbook = Workbook()
    _write_base_sheets(workbook)
    workbook["Industry Comments"].append(
        ["Construction", None, "Comment without a date"]
    )
    workbook.save(path)

    with pytest.raises(ValueError, match="no date"):
        import_ism_services.parse_industry_comments(path)
