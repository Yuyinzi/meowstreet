import json
import sys
from datetime import datetime
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

SECTOR_WORKBOOK = ROOT / "data" / "source_material" / "Video 23" / "US_Sector_Data.xlsm"
LONGS_WORKBOOK = ROOT / "data" / "source_material" / "Video 25" / "EG_Profiles_Longs.xlsx"
SHORTS_WORKBOOK = ROOT / "data" / "source_material" / "Video 26" / "EG_Profiles_Shorts.xlsx"
FIXTURES_DIR = ROOT / "tests" / "fixtures"

SAMPLE_TICKERS = ["DELL", "MYRG", "ARMK", "TRGP", "XMTR", "AAPL", "MSFT", "SHOP", "W", "OSTK"]

SECTOR_COLUMNS = {
    "symbol": 0,
    "company_name": 1,
    "naics_sector": 2,
    "naics_subsector": 3,
    "naics_industry": 4,
    "fiscal_year_end": 5,
    "market_cap": 6,
    "revenue_fy0": 7,
    "revenue_fy1": 8,
    "revenue_fy2": 9,
    "eps_fy0": 12,
    "eps_fy1": 13,
    "eps_fy2": 14,
    "eps_fy3": 15,
    "eg_f1": 16,
    "eg_f2": 17,
    "eg_f3": 18,
    "pe_fy1": 19,
    "pe_fy2": 20,
    "pe_fy3": 21,
    "peg_f1": 22,
    "peg_f2": 23,
    "peg_f3": 24,
}


def _cell_value(raw):
    if raw is None:
        return None
    if isinstance(raw, datetime):
        return raw.date().isoformat()
    text = str(raw).strip()
    if text in ("NULL", "NaN", ""):
        return None
    try:
        return float(text)
    except ValueError:
        return text


def _extract_sector_rows():
    wb = openpyxl.load_workbook(SECTOR_WORKBOOK, read_only=True)
    ws = wb["US Stock Screener >$1bn Mkt Cap"]
    rows_by_ticker = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        ticker = row[0]
        if ticker in SAMPLE_TICKERS and ticker not in rows_by_ticker:
            rows_by_ticker[ticker] = row
    wb.close()
    missing = [t for t in SAMPLE_TICKERS if t not in rows_by_ticker]
    if missing:
        raise ValueError(f"sample tickers missing from workbook: {missing}")
    extracted = []
    for ticker in SAMPLE_TICKERS:
        row = rows_by_ticker[ticker]
        record = {name: _cell_value(row[idx]) for name, idx in SECTOR_COLUMNS.items()}
        record["expected"] = {
            key: record.pop(key)
            for key in ("eg_f1", "eg_f2", "eg_f3", "pe_fy1", "pe_fy2", "pe_fy3", "peg_f1", "peg_f2", "peg_f3")
        }
        extracted.append(record)
    return extracted


def _extract_eg_profiles(path):
    wb = openpyxl.load_workbook(path)
    profiles = {}
    for ws in wb.worksheets:
        values = {c.coordinate: c.value for row in ws.iter_rows() for c in row if c.value is not None}
        annotations = sorted(
            (str(v) for coord, v in values.items() if isinstance(v, str) and coord not in ("A1", "B1", "C1", "A2", "A3")),
        )
        profiles[ws.title] = {
            "metric": values.get("B1"),
            "sector": {"v1": values.get("B2"), "v2": values.get("C2")},
            "stock": {"v1": values.get("B3"), "v2": values.get("C3")},
            "annotations": annotations,
        }
    wb.close()
    return profiles


def main():
    sector_rows = _extract_sector_rows()
    longs = _extract_eg_profiles(LONGS_WORKBOOK)
    shorts = _extract_eg_profiles(SHORTS_WORKBOOK)

    FIXTURES_DIR.mkdir(parents=True, exist_ok=True)
    sector_fixture = {
        "source": {
            "workbook": "data/source_material/Video 23/US_Sector_Data.xlsm",
            "sheet": "US Stock Screener >$1bn Mkt Cap",
            "workbook_created": "2020-04-13",
            "workbook_modified": "2021-10-26",
            "note": "2021-10 snapshot; parity fixtures only, never runtime data. "
            "Workbook stores raw formulas without the P28 sign-change override.",
        },
        "rows": sector_rows,
    }
    profiles_fixture = {
        "source": {
            "longs_workbook": "data/source_material/Video 25/EG_Profiles_Longs.xlsx",
            "shorts_workbook": "data/source_material/Video 26/EG_Profiles_Shorts.xlsx",
            "workbook_creator": "Anton Kreil",
            "note": "v1/v2 are EG1%/EG2% for EG profile sheets, PE1/PE2 for the PE Ideal sheet. "
            "Shorts case numbering differs from longs; shorts annotations reference the canonical longs cases.",
        },
        "longs": longs,
        "shorts": shorts,
    }

    sector_path = FIXTURES_DIR / "quant_screen_rows.json"
    profiles_path = FIXTURES_DIR / "eg_profiles.json"
    sector_path.write_text(json.dumps(sector_fixture, indent=2) + "\n")
    profiles_path.write_text(json.dumps(profiles_fixture, indent=2) + "\n")
    print(f"wrote {sector_path} ({len(sector_rows)} rows: {', '.join(SAMPLE_TICKERS)})")
    print(f"wrote {profiles_path} (longs sheets: {len(longs)}, shorts sheets: {len(shorts)})")


if __name__ == "__main__":
    main()
