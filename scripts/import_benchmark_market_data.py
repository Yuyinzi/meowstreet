import sys
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

from app.db import benchmark_market_data


DEFAULT_WORKBOOK_PATH = ROOT / "data" / "source_material" / "Video 02" / "Bull_Bear_Markets.xlsx"
WORKBOOK_BENCHMARK_SHEETS = [
    {"benchmark_id": "us_sp500", "sheet": "S&P 500"},
    {"benchmark_id": "us_nasdaq_100", "sheet": "Nasdaq 100"},
    {"benchmark_id": "us_nasdaq_composite", "sheet": "Nasdaq Composite"},
    {"benchmark_id": "us_djia", "sheet": "DJIA"},
    {"benchmark_id": "europe_stoxx_50", "sheet": "Eurostoxx 50"},
    {"benchmark_id": "europe_stoxx_600", "sheet": "Eurostoxx 600"},
    {"benchmark_id": "uk_ftse_100", "sheet": "FTSE 100"},
    {"benchmark_id": "uk_ftse_250", "sheet": "FTSE 250"},
    {"benchmark_id": "uk_ftse_350", "sheet": "FTSE 350"},
    {"benchmark_id": "germany_dax_40", "sheet": "DAX 40"},
    {"benchmark_id": "hong_kong_hsi", "sheet": "HSI"},
    {"benchmark_id": "hong_kong_hscei", "sheet": "HSCEI"},
    {"benchmark_id": "japan_nikkei_225", "sheet": "Nikkei 225"},
    {"benchmark_id": "australia_asx_200", "sheet": "ASX 200"},
]


def float_or_none(value):
    if value in (None, ""):
        return None
    return float(value)


def cell_date_iso(value):
    if hasattr(value, "date"):
        return value.date().isoformat()
    return str(value)


def load_workbook_sheet(workbook_path, sheet_name, data_only=True):
    path = Path(workbook_path)
    if not path.exists():
        raise ValueError(f"workbook does not exist: {path}")
    workbook = openpyxl.load_workbook(path, data_only=data_only, read_only=True)
    if sheet_name not in workbook.sheetnames:
        raise ValueError(f"workbook sheet is missing: {sheet_name}")
    return workbook[sheet_name]


def normalize_price_rows(raw_rows):
    return [
        {
            "date": str(row["date"]),
            "open": float_or_none(row.get("open")),
            "high": float_or_none(row.get("high")),
            "low": float_or_none(row.get("low")),
            "close": float(row["close"]),
        }
        for row in raw_rows
        if row.get("date") and row.get("close") not in (None, "")
    ]


def parse_workbook_sheet(workbook_path, sheet_name):
    sheet = load_workbook_sheet(workbook_path, sheet_name, data_only=False)
    raw_rows = []
    for values in sheet.iter_rows(min_row=2, values_only=True):
        if values[0] is None:
            continue
        if sheet_name == "S&P 500":
            date_value, open_value, high_value, low_value, close_value = values[:5]
        else:
            date_value = values[0]
            open_value = None
            high_value = values[1]
            low_value = None
            close_value = values[1]
        raw_rows.append(
            {
                "date": cell_date_iso(date_value),
                "open": open_value,
                "high": high_value if high_value is not None else close_value,
                "low": low_value,
                "close": close_value,
            }
        )
    return normalize_price_rows(reversed(raw_rows))


def import_price_rows(con, benchmark_id, rows, source):
    return benchmark_market_data.replace_benchmark_prices(
        con,
        benchmark_id,
        rows,
        source=source,
    )


def import_workbook(con, workbook_path=DEFAULT_WORKBOOK_PATH):
    inserted = {}
    errors = {}
    source = Path(workbook_path).name
    for config in WORKBOOK_BENCHMARK_SHEETS:
        try:
            rows = parse_workbook_sheet(workbook_path, config["sheet"])
            inserted[config["benchmark_id"]] = import_price_rows(
                con,
                config["benchmark_id"],
                rows,
                source=source,
            )
        except ValueError as exc:
            errors[config["benchmark_id"]] = str(exc)
    return inserted, errors


def main():
    con = benchmark_market_data.connect()
    inserted, errors = import_workbook(con)
    for benchmark_id, count in inserted.items():
        print(f"{benchmark_id}: {count}")
    for benchmark_id, message in errors.items():
        print(f"ERROR {benchmark_id}: {message}", file=sys.stderr)


if __name__ == "__main__":
    main()
