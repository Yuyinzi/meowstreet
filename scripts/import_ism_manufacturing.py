import argparse
import sys
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

from app.db import us_rates_liquidity

DEFAULT_WORKBOOK_PATH = (
    ROOT / "data" / "materials" / "Video 07" / "ISM_Manufacturing_Index.xlsx"
)

SERIES_CONFIG = {
    "ism_manufacturing_pmi": {
        "sheet": "PMI",
        "title": "ISM Manufacturing PMI",
        "units": "index",
    },
    "ism_manufacturing_new_orders": {
        "sheet": "New Orders",
        "title": "ISM New Orders",
        "units": "index",
    },
    "ism_manufacturing_production": {
        "sheet": "Production",
        "title": "ISM Production",
        "units": "index",
    },
    "ism_manufacturing_employment": {
        "sheet": "Employment",
        "title": "ISM Employment",
        "units": "index",
    },
    "ism_manufacturing_supplier_deliveries": {
        "sheet": "Deliveries",
        "title": "ISM Supplier Deliveries",
        "units": "index",
    },
    "ism_manufacturing_inventories": {
        "sheet": "Inventories",
        "title": "ISM Inventories",
        "units": "index",
    },
    "ism_manufacturing_customer_inventories": {
        "sheet": "Customer Inventories",
        "title": "ISM Customer Inventories",
        "units": "index",
    },
    "ism_manufacturing_prices": {
        "sheet": "Prices",
        "title": "ISM Prices",
        "units": "index",
    },
    "ism_manufacturing_order_backlog": {
        "sheet": "Order Backlog",
        "title": "ISM Order Backlog",
        "units": "index",
    },
    "ism_manufacturing_exports": {
        "sheet": "Exports",
        "title": "ISM Exports",
        "units": "index",
    },
    "ism_manufacturing_imports": {
        "sheet": "Imports",
        "title": "ISM Imports",
        "units": "index",
    },
}


def _iso_date(value):
    if isinstance(value, datetime):
        return value.date().isoformat()
    if hasattr(value, "isoformat"):
        return value.isoformat()
    return str(value)


SECTORS_SHEET = "Sectors"


def _sector_direction(value):
    if value == "Growth":
        return "growth"
    if value == "Contraction":
        return "contraction"
    return None


def _is_date(value):
    if isinstance(value, (datetime, date)):
        return True
    if isinstance(value, str):
        try:
            datetime.fromisoformat(value)
            return True
        except (ValueError, TypeError):
            return False
    return False


def parse_workbook(workbook_path=DEFAULT_WORKBOOK_PATH):
    workbook_path = Path(workbook_path)
    if not workbook_path.exists():
        raise ValueError(f"ism manufacturing workbook is missing: {workbook_path}")
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    results = []
    for series_id, config in SERIES_CONFIG.items():
        sheet_name = config["sheet"]
        if sheet_name not in workbook.sheetnames:
            raise ValueError(f"ism manufacturing sheet is missing: {sheet_name}")
        sheet = workbook[sheet_name]
        points = [
            {
                "date": _iso_date(date_value),
                "value": float(level_value),
                "source": workbook_path.name,
            }
            for date_value, level_value in sheet.iter_rows(
                min_row=2,
                min_col=1,
                max_col=2,
                values_only=True,
            )
            if date_value is not None
            and level_value not in (None, "")
            and _is_date(date_value)
        ]
        results.append(
            {
                "series": {
                    "series_id": series_id,
                    "title": config["title"],
                    "units": config["units"],
                    "source": workbook_path.name,
                },
                "points": points,
            }
        )
    return results


def parse_sector_rankings(workbook_path=DEFAULT_WORKBOOK_PATH):
    workbook_path = Path(workbook_path)
    if not workbook_path.exists():
        raise ValueError(f"ism manufacturing workbook is missing: {workbook_path}")
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    if SECTORS_SHEET not in workbook.sheetnames:
        raise ValueError(f"ism manufacturing sheet is missing: {SECTORS_SHEET}")
    sheet = workbook[SECTORS_SHEET]
    month_columns = [
        (column, _iso_date(sheet.cell(row=3, column=column).value))
        for column in range(3, sheet.max_column + 1, 2)
        if _is_date(sheet.cell(row=3, column=column).value)
    ]
    rows = []
    for row_index in range(6, sheet.max_row + 1):
        industry = sheet.cell(row=row_index, column=2).value
        if not industry:
            continue
        for status_column, month in month_columns:
            direction = _sector_direction(
                sheet.cell(row=row_index, column=status_column).value
            )
            rank = sheet.cell(row=row_index, column=status_column + 1).value
            if direction is None or rank in (None, ""):
                continue
            rows.append(
                {
                    "date": month,
                    "industry": str(industry),
                    "direction": direction,
                    "rank": int(rank),
                    "source": workbook_path.name,
                }
            )
    return rows


def import_workbook(con, workbook_path=DEFAULT_WORKBOOK_PATH):
    parsed_list = parse_workbook(workbook_path)
    results = {}
    for parsed in parsed_list:
        sid = parsed["series"]["series_id"]
        saved = us_rates_liquidity.replace_macro_indicator_points(
            con,
            parsed["series"],
            parsed["points"],
        )
        results[sid] = saved["points"]
    rankings = parse_sector_rankings(workbook_path)
    results["ism_industry_rankings"] = us_rates_liquidity.replace_ism_industry_rankings(
        con, rankings
    )
    return results


def main(argv=None):
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "--db-path", type=Path, default=us_rates_liquidity.DEFAULT_DB_PATH
    )
    parser.add_argument("--workbook-path", type=Path, default=DEFAULT_WORKBOOK_PATH)
    args = parser.parse_args(argv)
    con = us_rates_liquidity.connect(args.db_path)
    try:
        inserted = import_workbook(con, args.workbook_path)
    finally:
        con.close()
    for series_id, count in inserted.items():
        print(f"{series_id}: {count}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
