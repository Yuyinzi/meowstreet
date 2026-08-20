import argparse
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

from openpyxl import load_workbook

from app.data_sources import gics_reference
from app.db import ticker_context as ticker_context_db
from app.resources import resource_path


DEFAULT_REFERENCE_PATH = resource_path("gics_reference")


def workbook_tags(workbook_path):
    wb = load_workbook(workbook_path, data_only=True)
    ws = wb["Sheet1"]
    tags = {}
    for row in ws.iter_rows(min_row=4, max_row=87, max_col=7, values_only=True):
        _, _, _, industry, cyclical, defensive, both = row
        if industry is None:
            continue
        tag = (
            "cyclical"
            if cyclical
            else "defensive"
            if defensive
            else "both"
            if both
            else None
        )
        if tag is None:
            raise ValueError(f"workbook industry {industry} has no cycle tag")
        tags[str(industry).strip()] = tag
    return tags


def verify_against_workbook(industries, workbook_path):
    workbook = workbook_tags(workbook_path)
    reference = {row["industry"]: row["cycle_tag"] for row in industries}
    if reference != workbook:
        missing = sorted(set(workbook) - set(reference))
        extra = sorted(set(reference) - set(workbook))
        mismatched = sorted(
            name
            for name in set(reference) & set(workbook)
            if reference[name] != workbook[name]
        )
        raise ValueError(
            f"reference does not match the workbook: missing={missing} extra={extra} mismatched={mismatched}"
        )


def main(argv=None):
    parser = argparse.ArgumentParser(description="Import GICS industry cycle tags into SQLite")
    parser.add_argument("--reference", default=str(DEFAULT_REFERENCE_PATH))
    parser.add_argument("--db", default=str(ticker_context_db.DEFAULT_DB_PATH))
    parser.add_argument("--verify-workbook")
    args = parser.parse_args(argv)

    payload = gics_reference.load_gics_reference(args.reference)
    if args.verify_workbook is not None:
        verify_against_workbook(payload["industries"], args.verify_workbook)

    con = ticker_context_db.connect(args.db)
    try:
        saved_tags = ticker_context_db.save_industry_tags(con, payload["industries"])
        saved_aliases = ticker_context_db.save_industry_aliases(con, payload["aliases"])
    finally:
        con.close()

    print(f"imported {saved_tags} industry tags and {saved_aliases} yahoo aliases into {args.db}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
