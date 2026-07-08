import argparse
import sys
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

from app.data_sources.fred import FredClient
from app.data_sources.fred import parse_fred_csv
from app.db import us_rates_liquidity

DEFAULT_WORKBOOK_PATH = (
    ROOT / "data" / "materials" / "Video 06" / "US_M2_Money_Supply_Template.xlsx"
)
M2_SHEET_NAME = "Nominal M2 - Monthly"
M2_SERIES_ID = "m2_money_stock"
FRED_M2_SERIES_ID = "M2SL"
DEFAULT_FRED_DIR = DEFAULT_WORKBOOK_PATH.parent / "fred"
COMBINED_SOURCE = "P06 workbook + FRED"


def _iso_date(value):
    if isinstance(value, datetime):
        return value.date().isoformat()
    if hasattr(value, "isoformat"):
        return value.isoformat()
    return str(value)


def _series_payload(workbook_path):
    return {
        "series_id": M2_SERIES_ID,
        "title": "M2 Money Stock",
        "units": "billions_usd",
        "source": Path(workbook_path).name,
    }


def _point_payload(date_value, level_value, workbook_path):
    return {
        "date": _iso_date(date_value),
        "value": float(level_value),
        "source": Path(workbook_path).name,
    }


def _fred_series_payload():
    return {
        "series_id": M2_SERIES_ID,
        "title": "M2 Money Stock",
        "units": "billions_usd",
        "source": COMBINED_SOURCE,
    }


def _fred_points_payload(rows, csv_path):
    return [
        {
            "date": date_key,
            "value": value,
            "source": Path(csv_path).name,
        }
        for date_key, value in rows.items()
    ]


def build_fred_m2_payload(csv_path):
    rows = parse_fred_csv(csv_path, FRED_M2_SERIES_ID)
    return {
        "series": _fred_series_payload(),
        "points": _fred_points_payload(rows, csv_path),
    }


def parse_workbook(workbook_path=DEFAULT_WORKBOOK_PATH):
    workbook_path = Path(workbook_path)
    if not workbook_path.exists():
        raise ValueError(f"m2 money supply workbook is missing: {workbook_path}")
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    if M2_SHEET_NAME not in workbook.sheetnames:
        raise ValueError(f"m2 money supply sheet is missing: {M2_SHEET_NAME}")
    sheet = workbook[M2_SHEET_NAME]
    points = [
        _point_payload(date_value, level_value, workbook_path)
        for date_value, level_value in sheet.iter_rows(
            min_row=2,
            min_col=1,
            max_col=2,
            values_only=True,
        )
        if date_value is not None and level_value not in (None, "")
    ]
    return {"series": _series_payload(workbook_path), "points": points}


def import_workbook(con, workbook_path=DEFAULT_WORKBOOK_PATH):
    payload = parse_workbook(workbook_path)
    saved = us_rates_liquidity.replace_macro_indicator_points(
        con,
        payload["series"],
        payload["points"],
    )
    return {payload["series"]["series_id"]: saved["points"]}


def import_fred_csvs(con, fred_dir=DEFAULT_FRED_DIR):
    payload = build_fred_m2_payload(Path(fred_dir) / f"{FRED_M2_SERIES_ID}.csv")
    saved = us_rates_liquidity.merge_macro_indicator_points(
        con,
        payload["series"],
        payload["points"],
    )
    return {payload["series"]["series_id"]: saved["points"]}


def fetch_fred_csvs(fred_dir=DEFAULT_FRED_DIR):
    client = FredClient(fred_dir)
    return client.fetch_csvs([FRED_M2_SERIES_ID])


def main(argv=None):
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "--db-path", type=Path, default=us_rates_liquidity.DEFAULT_DB_PATH
    )
    parser.add_argument("--workbook-path", type=Path, default=DEFAULT_WORKBOOK_PATH)
    parser.add_argument("--fred-dir", type=Path, default=DEFAULT_FRED_DIR)
    parser.add_argument("--fetch-fred-csv", action="store_true")
    parser.add_argument("--fred-csv-merge", action="store_true")
    args = parser.parse_args(argv)
    if args.fetch_fred_csv:
        fetched = fetch_fred_csvs(args.fred_dir)
        for series_id, path in fetched.items():
            print(f"{series_id}: {path}")
        return 0
    con = us_rates_liquidity.connect(args.db_path)
    try:
        if args.fred_csv_merge:
            inserted = import_fred_csvs(con, args.fred_dir)
        else:
            inserted = import_workbook(con, args.workbook_path)
    finally:
        con.close()
    for series_id, count in inserted.items():
        print(f"{series_id}: {count}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
